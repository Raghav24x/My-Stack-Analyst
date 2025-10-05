# -*- coding: utf-8 -*-
import requests
import json
import re
import time
from datetime import datetime, timedelta
from typing import Dict, List, Optional
from bs4 import BeautifulSoup
from urllib.parse import urljoin
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

class SubstackDataCollector:
    def __init__(self, publication_input: str, api_key: Optional[str] = None):
        self.api_key = api_key
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            'Cache-Control': 'no-cache, no-store, must-revalidate',
            'Pragma': 'no-cache',
            'Expires': '0'
        })
        
        # Handle different input formats
        if publication_input.startswith('http'):
            # Full URL provided
            if publication_input.endswith('/'):
                publication_input = publication_input[:-1]
            if publication_input.endswith('/feed'):
                self.base_url = publication_input.replace('/feed', '')
            else:
                self.base_url = publication_input
            # Extract publication name from URL
            self.publication_name = self.base_url.split('//')[-1].split('.')[0]
        else:
            # Just publication name provided
            self.publication_name = publication_input
            self.base_url = f"https://{publication_input}.substack.com"
    
    def fetch_posts(self, limit: int = None) -> List[Dict]:
        """Fetch posts from the publication RSS feed."""
        try:
            url = f"{self.base_url}/feed"
            response = requests.get(url)
            response.raise_for_status()
            posts = self._parse_rss_feed(response.text)
            
            # If limit is specified, return only that many posts
            if limit is not None:
                return posts[:limit]
            else:
                # Return all posts
                return posts
        except Exception as e:
            print(f"Error fetching posts: {e}")
            return []
    
    def _parse_rss_feed(self, rss_content: str) -> List[Dict]:
        """Parse RSS feed content to extract post data."""
        from bs4 import BeautifulSoup
        
        soup = BeautifulSoup(rss_content, "xml")
        posts = []
        
        for item in soup.find_all("item"):
            post = {
                "title": item.find("title").text if item.find("title") else "",
                "link": item.find("link").text if item.find("link") else "",
                "description": item.find("description").text if item.find("description") else "",
                "pub_date": item.find("pubDate").text if item.find("pubDate") else "",
                "author": item.find("author").text if item.find("author") else "",
            }
            posts.append(post)
        
        return posts
    
    def get_publication_info(self) -> Dict:
        """Get basic publication information using HTML scraping."""
        try:
            print(f"[DEBUG] Fetching publication page: {self.base_url}")
            # Add cache-busting parameter to get fresh data
            import time
            cache_buster = int(time.time() * 1000)
            url_with_cache_buster = f"{self.base_url}?t={cache_buster}"
            
            response = self.session.get(url_with_cache_buster)
            response.raise_for_status()
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Save HTML for debugging (optional)
            debug_filename = f"debug_{self.publication_name}_page_{cache_buster}.html"
            with open(debug_filename, 'w', encoding='utf-8') as f:
                f.write(response.text)
            print(f"[DEBUG] Saved HTML content to {debug_filename}")
            
            # Extract publication details
            title = soup.find('title')
            title_text = title.text if title else f"{self.publication_name} | Substack"
            
            # Try to find subscriber count with comprehensive scraping
            subscriber_count = self._extract_subscriber_count(soup)
            
            # If subscriber count not found, try alternative methods
            if subscriber_count is None:
                print(f"[INFO] Subscriber count not found in HTML. Trying alternative methods...")
                # Try the API method as fallback
                subscriber_count = self._get_subscriber_count_from_api()
                
            if subscriber_count is None:
                print(f"[INFO] Subscriber count not publicly available for this publication")
            
            return {
                'name': title_text,
                'url': self.base_url,
                'subscriber_count': subscriber_count,
                'last_updated': datetime.now().isoformat()
            }
        except Exception as e:
            print(f"Error fetching publication info: {e}")
            return {
                'name': f"{self.publication_name} | Substack",
                'url': self.base_url,
                'subscriber_count': None,
                'last_updated': datetime.now().isoformat()
            }
    
    def _get_subscriber_count_from_api(self) -> Optional[int]:
        """Get subscriber count from Substack's public search API."""
        try:
            api_url = "https://substack.com/api/v1/publication/search"
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
                'Accept': 'application/json',
                'Origin': 'https://substack.com',
                'Referer': 'https://substack.com/discover'
            }
            params = {
                'query': self.publication_name,
                'limit': 20  # Get more results to find the right publication
            }
            
            print(f"[DEBUG] Querying Substack API for publication: {self.publication_name}")
            response = requests.get(api_url, headers=headers, params=params)
            response.raise_for_status()
            data = response.json()
            
            # Find the matching publication
            if 'publications' in data and data['publications']:
                for pub in data['publications']:
                    # Match by subdomain or name (more flexible matching)
                    subdomain = pub.get('subdomain', '').lower()
                    pub_name = pub.get('name', '').lower()
                    pub_url = pub.get('url', '').lower()
                    
                    # More flexible matching
                    if (subdomain == self.publication_name.lower() or 
                        self.publication_name.lower() in pub_name or
                        self.publication_name.lower() in pub_url or
                        self.base_url.lower() in pub_url):
                        
                        subscriber_count = pub.get('subscriber_count')
                        if subscriber_count is not None and subscriber_count > 0:
                            print(f"[SUCCESS] Found subscriber count from API: {subscriber_count}")
                            return int(subscriber_count)
                        else:
                            print(f"[INFO] Found publication but subscriber_count is null or zero")
                            continue
                
                print(f"[INFO] Publication '{self.publication_name}' not found in API results")
            else:
                print(f"[INFO] No publications found in API response")
            
            return None
        except Exception as e:
            print(f"[ERROR] Error fetching subscriber count from API: {e}")
            return None
    
    def _extract_subscriber_count_from_js(self, html_content: str) -> Optional[int]:
        """Extract subscriber count from JavaScript data in the HTML."""
        try:
            import re
            import json
            
            # Look for the window.__NEXT_DATA__ or similar JavaScript data
            # Pattern to find subscriber count in the JSON data
            patterns = [
                r'"subscriber_count":\s*(\d+)',
                r'"subscribers":\s*(\d+)',
                r'"total_subscribers":\s*(\d+)',
                r'"subscriberCount":\s*(\d+)',
                r'"subscriber_count":\s*"(\d+)"',
                r'"subscribers":\s*"(\d+)"',
                r'"total_subscribers":\s*"(\d+)"',
                r'"subscriberCount":\s*"(\d+)"'
            ]
            
            for pattern in patterns:
                matches = re.findall(pattern, html_content, re.IGNORECASE)
                if matches:
                    # Get the first match and convert to int
                    count = int(matches[0])
                    print(f"[DEBUG] Found subscriber count: {count}")
                    return count
            
            # Try to find in window.__NEXT_DATA__ specifically
            next_data_match = re.search(r'window\.__NEXT_DATA__\s*=\s*({.*?});', html_content, re.DOTALL)
            if next_data_match:
                try:
                    next_data = json.loads(next_data_match.group(1))
                    # Navigate through the JSON structure to find subscriber count
                    if 'props' in next_data and 'pageProps' in next_data['props']:
                        page_props = next_data['props']['pageProps']
                        if 'publication' in page_props and 'subscriber_count' in page_props['publication']:
                            count = page_props['publication']['subscriber_count']
                            print(f"[DEBUG] Found subscriber count in __NEXT_DATA__: {count}")
                            return int(count)
                except (json.JSONDecodeError, KeyError, ValueError) as e:
                    print(f"[DEBUG] Error parsing __NEXT_DATA__: {e}")
                    pass
            
            print("[DEBUG] No subscriber count found in JavaScript data - this is normal for many Substack publications")
            print("[DEBUG] Substack often doesn't expose subscriber counts publicly for privacy reasons")
            return None
            
        except Exception as e:
            print(f"[DEBUG] Error extracting subscriber count from JS: {e}")
            return None
    
    def _extract_subscriber_count(self, soup: BeautifulSoup) -> Optional[int]:
        """Extract subscriber count from the publication page with comprehensive scraping."""
        subscriber_count = None
        
        print(f"[DEBUG] Searching for subscriber count in {self.base_url}")
        
        # Method 1: Look for Substack-specific subscriber elements
        # Common Substack patterns for subscriber count
        subscriber_patterns = [
            r'(\d+)\s*subscribers?',
            r'(\d+)\s*readers?',
            r'(\d+)\s*members?',
            r'(\d+)\s*people\s+subscribed',
            r'(\d+)\s*subscribed'
        ]
        
        # Search in all text content
        page_text = soup.get_text()
        for pattern in subscriber_patterns:
            matches = re.findall(pattern, page_text, re.IGNORECASE)
            for match in matches:
                try:
                    count = int(match)
                    if count > 0 and count < 1000000:  # Reasonable range
                        subscriber_count = count
                        print(f"[DEBUG] Found subscriber count via text pattern: {count}")
                        break
                except ValueError:
                    continue
            if subscriber_count:
                break
        
        # Method 2: Look for specific HTML elements with subscriber data
        if not subscriber_count:
            # Look for elements with data attributes
            data_elements = soup.find_all(attrs={'data-subscriber-count': True})
            for element in data_elements:
                try:
                    count = int(element.get('data-subscriber-count'))
                    if count > 0:
                        subscriber_count = count
                        print(f"[DEBUG] Found subscriber count via data attribute: {count}")
                        break
                except (ValueError, TypeError):
                    continue
        
        # Method 3: Look for specific CSS classes that might contain subscriber count
        if not subscriber_count:
            class_patterns = [
                'subscriber-count', 'subscriberCount', 'subscribers',
                'reader-count', 'readerCount', 'readers',
                'member-count', 'memberCount', 'members'
            ]
            
            for class_pattern in class_patterns:
                elements = soup.find_all(class_=re.compile(class_pattern, re.IGNORECASE))
                for element in elements:
                    text = element.get_text()
                    match = re.search(r'(\d+)', text)
                    if match:
                        try:
                            count = int(match.group(1))
                            if count > 0 and count < 1000000:
                                subscriber_count = count
                                print(f"[DEBUG] Found subscriber count via class {class_pattern}: {count}")
                                break
                        except ValueError:
                            continue
                if subscriber_count:
                    break
        
        # Method 4: Look for JavaScript variables or data
        if not subscriber_count:
            script_tags = soup.find_all('script')
            for script in script_tags:
                if script.string:
                    # Look for subscriber count in JavaScript variables
                    js_patterns = [
                        r'subscriberCount["\']?\s*[:=]\s*(\d+)',
                        r'subscribers["\']?\s*[:=]\s*(\d+)',
                        r'readerCount["\']?\s*[:=]\s*(\d+)',
                        r'memberCount["\']?\s*[:=]\s*(\d+)',
                        r'"subscriber_count"\s*:\s*(\d+)',
                        r'"subscribers"\s*:\s*(\d+)'
                    ]
                    
                    for pattern in js_patterns:
                        matches = re.findall(pattern, script.string, re.IGNORECASE)
                        for match in matches:
                            try:
                                count = int(match)
                                if count > 0 and count < 1000000:
                                    subscriber_count = count
                                    print(f"[DEBUG] Found subscriber count via JavaScript: {count}")
                                    break
                            except ValueError:
                                continue
                        if subscriber_count:
                            break
                    if subscriber_count:
                        break
        
        # Method 5: Look for meta tags with subscriber information
        if not subscriber_count:
            meta_tags = soup.find_all('meta')
            for meta in meta_tags:
                name = meta.get('name', '').lower()
                property_attr = meta.get('property', '').lower()
                content = meta.get('content', '')
                
                if any(keyword in name or keyword in property_attr for keyword in ['subscriber', 'reader', 'member']):
                    match = re.search(r'(\d+)', content)
                    if match:
                        try:
                            count = int(match.group(1))
                            if count > 0 and count < 1000000:
                                subscriber_count = count
                                print(f"[DEBUG] Found subscriber count via meta tag: {count}")
                                break
                        except ValueError:
                            continue
        
        # Method 6: Look for JSON-LD structured data
        if not subscriber_count:
            json_scripts = soup.find_all('script', type='application/ld+json')
            for script in json_scripts:
                try:
                    data = json.loads(script.string)
                    if isinstance(data, dict):
                        # Look for subscriber count in various possible keys
                        for key in ['subscriberCount', 'subscribers', 'memberCount', 'readers', 'subscriber_count']:
                            if key in data and isinstance(data[key], (int, str)):
                                try:
                                    count = int(data[key])
                                    if count > 0 and count < 1000000:
                                        subscriber_count = count
                                        print(f"[DEBUG] Found subscriber count via JSON-LD: {count}")
                                        break
                                except ValueError:
                                    continue
                        if subscriber_count:
                            break
                except (json.JSONDecodeError, TypeError):
                    continue
        
        # Method 7: Look for specific Substack UI elements
        if not subscriber_count:
            # Look for common Substack UI patterns
            ui_selectors = [
                '[data-testid*="subscriber"]',
                '[data-testid*="reader"]',
                '[data-testid*="member"]',
                '.subscriber-count',
                '.reader-count',
                '.member-count'
            ]
            
            for selector in ui_selectors:
                elements = soup.select(selector)
                for element in elements:
                    text = element.get_text()
                    match = re.search(r'(\d+)', text)
                    if match:
                        try:
                            count = int(match.group(1))
                            if count > 0 and count < 1000000:
                                subscriber_count = count
                                print(f"[DEBUG] Found subscriber count via UI selector {selector}: {count}")
                                break
                        except ValueError:
                            continue
                    if subscriber_count:
                        break
                if subscriber_count:
                    break
        
        # Method 8: Try to extract from JSON data in the HTML
        if not subscriber_count:
            subscriber_count = self._extract_subscriber_count_from_json(str(soup))
            if subscriber_count:
                print(f"[DEBUG] Found subscriber count via JSON data: {subscriber_count}")
        
        if subscriber_count:
            print(f"[DEBUG] Final subscriber count: {subscriber_count}")
        else:
            print(f"[DEBUG] No subscriber count found")
        
        return subscriber_count
    
    def _extract_subscriber_count_from_json(self, html_content: str) -> Optional[int]:
        """Extract subscriber count from JSON data in the HTML."""
        try:
            import re
            import json
            
            # First, try publication-specific patterns for any publication
            pub_patterns = [
                rf'"publication":\s*{{[^}}]*"subdomain":\s*"{self.publication_name}"[^}}]*"subscriber_count":\s*(\d+)',
                rf'"publication":\s*{{[^}}]*"subdomain":\s*"{self.publication_name}"[^}}]*"subscribers":\s*(\d+)',
                rf'"publication":\s*{{[^}}]*"subdomain":\s*"{self.publication_name}"[^}}]*"total_subscribers":\s*(\d+)',
                rf'"publication":\s*{{[^}}]*"subdomain":\s*"{self.publication_name}"[^}}]*"subscriberCount":\s*(\d+)',
                rf'"publication":\s*{{[^}}]*"subdomain":\s*"{self.publication_name}"[^}}]*"subscriber_count":\s*"(\d+)"',
                rf'"publication":\s*{{[^}}]*"subdomain":\s*"{self.publication_name}"[^}}]*"subscribers":\s*"(\d+)"',
                rf'"publication":\s*{{[^}}]*"subdomain":\s*"{self.publication_name}"[^}}]*"total_subscribers":\s*"(\d+)"',
                rf'"publication":\s*{{[^}}]*"subdomain":\s*"{self.publication_name}"[^}}]*"subscriberCount":\s*"(\d+)"'
            ]
            
            for pattern in pub_patterns:
                matches = re.findall(pattern, html_content, re.IGNORECASE | re.DOTALL)
                for match in matches:
                    try:
                        count = int(match)
                        if count > 0 and count < 1000000:
                            print(f"[DEBUG] Found subscriber count via JSON pattern for {self.publication_name}: {count}")
                            return count
                    except ValueError:
                        continue
            
            # Fallback to general patterns
            general_patterns = [
                r'"subscriber_count":\s*(\d+)',
                r'"subscribers":\s*(\d+)',
                r'"total_subscribers":\s*(\d+)',
                r'"subscriberCount":\s*(\d+)',
                r'"subscriber_count":\s*"(\d+)"',
                r'"subscribers":\s*"(\d+)"',
                r'"total_subscribers":\s*"(\d+)"',
                r'"subscriberCount":\s*"(\d+)"'
            ]
            
            for pattern in general_patterns:
                matches = re.findall(pattern, html_content, re.IGNORECASE)
                for match in matches:
                    try:
                        count = int(match)
                        if count > 0 and count < 1000000:
                            print(f"[DEBUG] Found subscriber count via general JSON pattern: {count}")
                            return count
                    except ValueError:
                        continue
            
            print(f"[DEBUG] No subscriber count found in JSON data.")
            return None
        except Exception as e:
            print(f"Error extracting subscriber count from JSON: {e}")
            return None
    
    def get_post_engagement(self, post_url: str) -> Dict:
        """Get engagement metrics for a specific post."""
        try:
            response = self.session.get(post_url)
            response.raise_for_status()
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Extract engagement metrics with improved accuracy
            likes_str = self._extract_likes(soup)
            comments_str = self._extract_comments(soup)
            shares_str = self._extract_shares(soup)
            restacks_str = self._extract_restacks(soup)
            
            # Convert to numbers or keep as "-"
            likes = int(likes_str) if likes_str != "-" else 0
            comments = int(comments_str) if comments_str != "-" else 0
            shares = int(shares_str) if shares_str != "-" else 0
            restacks = int(restacks_str) if restacks_str != "-" else 0
            
            # Extract word count from post content
            content = soup.find('div', class_='post-content') or soup.find('article')
            word_count = len(content.get_text().split()) if content else 0
            
            # Calculate reading time (average 200 words per minute)
            reading_time = max(1, word_count // 200) if word_count > 0 else 0
            
            return {
                'likes': likes_str,  # Keep original string for display
                'comments': comments_str,
                'shares': shares_str,
                'restacks': restacks_str,
                'likes_num': likes,  # Numeric version for calculations
                'comments_num': comments,
                'shares_num': shares,
                'restacks_num': restacks,
                'word_count': word_count,
                'reading_time': reading_time,
                'total_engagement': likes + comments + shares + restacks
            }
        except Exception as e:
            print(f"Error fetching engagement for {post_url}: {e}")
            return {
                'likes': "-",
                'comments': "-",
                'shares': "-",
                'restacks': "-",
                'likes_num': 0,
                'comments_num': 0,
                'shares_num': 0,
                'restacks_num': 0,
                'word_count': 0,
                'reading_time': 0,
                'total_engagement': 0
            }
    
    def _extract_likes(self, soup: BeautifulSoup) -> str:
        """Extract likes count with improved accuracy."""
        # Look for Substack-specific like elements
        like_selectors = [
            'button[data-testid="like-button"]',
            '.like-button',
            '[class*="like"]',
            'button[aria-label*="like"]',
            'button[title*="like"]'
        ]
        
        for selector in like_selectors:
            elements = soup.select(selector)
            for element in elements:
                # Look for number in the element or its children
                text = element.get_text()
                match = re.search(r'(\d+)', text)
                if match:
                    try:
                        number = int(match.group(1))
                        if number <= 1000:  # Realistic cap
                            return str(number)
                    except ValueError:
                        continue
        
        # Fallback to general text search
        text_content = soup.get_text()
        patterns = [r'(\d+)\s*likes?', r'(\d+)\s*hearts?', r'(\d+)\s*favorites?']
        for pattern in patterns:
            match = re.search(pattern, text_content, re.IGNORECASE)
            if match:
                try:
                    number = int(match.group(1))
                    if number <= 1000:  # Realistic cap
                        return str(number)
                except ValueError:
                    continue
        
        return "-"
    
    def _extract_comments(self, soup: BeautifulSoup) -> str:
        """Extract comments count with improved accuracy."""
        # Look for Substack-specific comment elements
        comment_selectors = [
            'button[data-testid="comment-button"]',
            '.comment-button',
            '[class*="comment"]',
            'button[aria-label*="comment"]',
            'button[title*="comment"]'
        ]
        
        for selector in comment_selectors:
            elements = soup.select(selector)
            for element in elements:
                text = element.get_text()
                match = re.search(r'(\d+)', text)
                if match:
                    try:
                        number = int(match.group(1))
                        if number <= 100:  # Realistic cap
                            return str(number)
                    except ValueError:
                        continue
        
        # Fallback to general text search
        text_content = soup.get_text()
        patterns = [r'(\d+)\s*comments?', r'(\d+)\s*replies?']
        for pattern in patterns:
            match = re.search(pattern, text_content, re.IGNORECASE)
            if match:
                try:
                    number = int(match.group(1))
                    if number <= 100:  # Realistic cap
                        return str(number)
                except ValueError:
                    continue
        
        return "-"
    
    def _extract_shares(self, soup: BeautifulSoup) -> str:
        """Extract shares count with improved accuracy."""
        # Look for Substack-specific share elements
        share_selectors = [
            'button[data-testid="share-button"]',
            '.share-button',
            '[class*="share"]',
            'button[aria-label*="share"]',
            'button[title*="share"]'
        ]
        
        for selector in share_selectors:
            elements = soup.select(selector)
            for element in elements:
                text = element.get_text()
                match = re.search(r'(\d+)', text)
                if match:
                    try:
                        number = int(match.group(1))
                        if number <= 1000:  # Realistic cap
                            return str(number)
                    except ValueError:
                        continue
        
        # Fallback to general text search
        text_content = soup.get_text()
        patterns = [r'(\d+)\s*shares?', r'(\d+)\s*retweets?']
        for pattern in patterns:
            match = re.search(pattern, text_content, re.IGNORECASE)
            if match:
                try:
                    number = int(match.group(1))
                    if number <= 1000:  # Realistic cap
                        return str(number)
                except ValueError:
                    continue
        
        return "-"
    
    def _extract_restacks(self, soup: BeautifulSoup) -> str:
        """Extract restacks count with improved accuracy."""
        # Look for Substack-specific restack elements
        restack_selectors = [
            'button[data-testid="restack-button"]',
            '.restack-button',
            '[class*="restack"]',
            'button[aria-label*="restack"]',
            'button[title*="restack"]',
            'button[aria-label*="repost"]',
            'button[title*="repost"]'
        ]
        
        for selector in restack_selectors:
            elements = soup.select(selector)
            for element in elements:
                text = element.get_text()
                match = re.search(r'(\d+)', text)
                if match:
                    try:
                        number = int(match.group(1))
                        if number <= 1000:  # Realistic cap
                            return str(number)
                    except ValueError:
                        continue
        
        # Fallback to general text search for restack patterns
        text_content = soup.get_text()
        patterns = [
            r'(\d+)\s*restacks?',
            r'(\d+)\s*reposts?',
            r'(\d+)\s*re-shares?',
            r'(\d+)\s*re-shares?'
        ]
        for pattern in patterns:
            match = re.search(pattern, text_content, re.IGNORECASE)
            if match:
                try:
                    number = int(match.group(1))
                    if number <= 1000:  # Realistic cap
                        return str(number)
                except ValueError:
                    continue
        
        return "-"
    
    def _extract_number(self, soup: BeautifulSoup, keywords: List[str]) -> str:
        """Legacy method - kept for compatibility."""
        text_content = soup.get_text()
        for keyword in keywords:
            pattern = rf'(\d+)\s*{keyword}s?'
            match = re.search(pattern, text_content, re.IGNORECASE)
            if match:
                try:
                    number = int(match.group(1))
                    if keyword in ['share', 'retweet'] and number > 1000:
                        return str(min(number, 1000))
                    elif keyword in ['like', 'heart', 'favorite'] and number > 1000:
                        return str(min(number, 500))
                    elif keyword in ['comment', 'reply'] and number > 100:
                        return str(min(number, 50))
                    return str(number)
                except ValueError:
                    continue
        return "-"
    
    def analyze_publication(self, limit: int = None) -> Dict:
        """Perform comprehensive analysis of the publication."""
        print("[ANALYSIS] Starting comprehensive analysis...")
        
        # Get publication info
        print("[ANALYSIS] Fetching publication information...")
        pub_info = self.get_publication_info()
        
        # Get posts - if no limit specified, get all posts
        print("[ANALYSIS] Fetching posts from RSS feed...")
        posts = self.fetch_posts(limit) if limit else self.fetch_posts()
        
        if not posts:
            return {"error": "No posts found"}
        
        total_posts = len(posts)
        print(f"[ANALYSIS] Found {total_posts} posts to analyze")
        
        # Analyze each post
        print("[ANALYSIS] Analyzing post engagement...")
        analyzed_posts = []
        total_likes = 0
        total_comments = 0
        total_shares = 0
        total_restacks = 0
        total_words = 0
        total_reading_time = 0
        
        for i, post in enumerate(posts):  # Analyze ALL posts
            # Safe encoding for display
            safe_title = post['title'].encode('ascii', 'ignore').decode('ascii')
            print(f"   Analyzing post {i+1}/{total_posts}: {safe_title[:50]}...")
            
            # Get engagement data
            engagement = self.get_post_engagement(post['link'])
            
            # Combine post data with engagement
            analyzed_post = {
                **post,
                **engagement,
                'engagement_rate': 0  # Will calculate later
            }
            analyzed_posts.append(analyzed_post)
            
            # Accumulate totals
            total_likes += engagement['likes_num']
            total_comments += engagement['comments_num']
            total_shares += engagement['shares_num']
            total_restacks += engagement['restacks_num']
            total_words += engagement['word_count']
            total_reading_time += engagement['reading_time']
            
            # Be respectful with requests
            time.sleep(1)
        
        # Calculate analytics
        num_posts = len(analyzed_posts)
        avg_likes = total_likes / num_posts if num_posts > 0 else 0
        avg_comments = total_comments / num_posts if num_posts > 0 else 0
        avg_shares = total_shares / num_posts if num_posts > 0 else 0
        avg_restacks = total_restacks / num_posts if num_posts > 0 else 0
        avg_words = total_words / num_posts if num_posts > 0 else 0
        avg_reading_time = total_reading_time / num_posts if num_posts > 0 else 0
        
        # Find top performing posts
        top_posts = sorted(analyzed_posts, key=lambda x: x['total_engagement'], reverse=True)[:5]
        
        # Calculate publishing frequency
        if len(posts) > 1:
            dates = [datetime.strptime(post['pub_date'], '%a, %d %b %Y %H:%M:%S %Z') for post in posts if post['pub_date']]
            if len(dates) > 1:
                date_range = max(dates) - min(dates)
                posts_per_week = len(posts) / (date_range.days / 7) if date_range.days > 0 else 0
            else:
                posts_per_week = 0
        else:
            posts_per_week = 0
        
        return {
            'publication': pub_info,
            'analytics': {
                'total_posts_analyzed': num_posts,
                'average_likes_per_post': round(avg_likes, 1),
                'average_comments_per_post': round(avg_comments, 1),
                'average_shares_per_post': round(avg_shares, 1),
                'average_restacks_per_post': round(avg_restacks, 1),
                'average_word_count': round(avg_words, 0),
                'average_reading_time': round(avg_reading_time, 1),
                'publishing_frequency': round(posts_per_week, 1),
                'total_engagement': total_likes + total_comments + total_shares + total_restacks
            },
            'top_posts': top_posts,
            'all_posts': analyzed_posts
        }
    
    def export_to_excel(self, analysis_data: Dict, filename: str = None) -> str:
        """Export analytics data to Excel file."""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"substack_analytics_{self.publication_name}_{timestamp}.xlsx"
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        overview_sheet = wb.create_sheet("Publication Overview")
        analytics_sheet = wb.create_sheet("Analytics Summary")
        posts_sheet = wb.create_sheet("All Posts")
        top_posts_sheet = wb.create_sheet("Top Posts")
        
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # 1. Publication Overview Sheet
        overview_sheet['A1'] = "Publication Overview"
        overview_sheet['A1'].font = Font(bold=True, size=16)
        
        pub = analysis_data['publication']
        overview_data = [
            ["Property", "Value"],
            ["Publication Name", pub['name']],
            ["URL", pub['url']],
            ["Subscriber Count", pub['subscriber_count'] or "Not publicly available (privacy setting)"],
            ["Last Updated", pub['last_updated']],
            ["Analysis Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        ]
        
        for row_idx, (prop, value) in enumerate(overview_data, 1):
            overview_sheet[f'A{row_idx}'] = prop
            overview_sheet[f'B{row_idx}'] = value
            if row_idx == 1:  # Header
                overview_sheet[f'A{row_idx}'].font = header_font
                overview_sheet[f'A{row_idx}'].fill = header_fill
                overview_sheet[f'B{row_idx}'].font = header_font
                overview_sheet[f'B{row_idx}'].fill = header_fill
        
        # 2. Analytics Summary Sheet
        analytics_sheet['A1'] = "Performance Metrics"
        analytics_sheet['A1'].font = Font(bold=True, size=16)
        
        analytics = analysis_data['analytics']
        analytics_data = [
            ["Metric", "Value"],
            ["Posts Analyzed", analytics['total_posts_analyzed']],
            ["Average Likes per Post", analytics['average_likes_per_post']],
            ["Average Comments per Post", analytics['average_comments_per_post']],
            ["Average Shares per Post", analytics['average_shares_per_post']],
            ["Average Restacks per Post", analytics['average_restacks_per_post']],
            ["Average Word Count", analytics['average_word_count']],
            ["Average Reading Time (minutes)", analytics['average_reading_time']],
            ["Publishing Frequency (posts/week)", analytics['publishing_frequency']],
            ["Total Engagement", analytics['total_engagement']]
        ]
        
        for row_idx, (metric, value) in enumerate(analytics_data, 1):
            analytics_sheet[f'A{row_idx}'] = metric
            analytics_sheet[f'B{row_idx}'] = value
            if row_idx == 1:  # Header
                analytics_sheet[f'A{row_idx}'].font = header_font
                analytics_sheet[f'A{row_idx}'].fill = header_fill
                analytics_sheet[f'B{row_idx}'].font = header_font
                analytics_sheet[f'B{row_idx}'].fill = header_fill
        
        # 3. All Posts Sheet
        posts_headers = ["Title", "Published Date", "Likes", "Comments", "Shares", "Restacks",
                        "Word Count", "Reading Time (min)", "Total Engagement", "Link"]
        for col_idx, header in enumerate(posts_headers, 1):
            cell = posts_sheet.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        for row_idx, post in enumerate(analysis_data['all_posts'], 2):
            posts_sheet[f'A{row_idx}'] = post['title']
            posts_sheet[f'B{row_idx}'] = post['pub_date']
            posts_sheet[f'C{row_idx}'] = post['likes']
            posts_sheet[f'D{row_idx}'] = post['comments']
            posts_sheet[f'E{row_idx}'] = post['shares']
            posts_sheet[f'F{row_idx}'] = post['restacks']
            posts_sheet[f'G{row_idx}'] = post['word_count']
            posts_sheet[f'H{row_idx}'] = post['reading_time']
            posts_sheet[f'I{row_idx}'] = post['total_engagement']
            posts_sheet[f'J{row_idx}'] = post['link']
        
        # 4. Top Posts Sheet
        top_headers = ["Rank", "Title", "Total Engagement", "Likes", "Comments", "Shares", "Restacks",
                      "Word Count", "Reading Time (min)", "Published Date"]
        for col_idx, header in enumerate(top_headers, 1):
            cell = top_posts_sheet.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        for row_idx, post in enumerate(analysis_data['top_posts'], 2):
            top_posts_sheet[f'A{row_idx}'] = row_idx - 1  # Rank
            top_posts_sheet[f'B{row_idx}'] = post['title']
            top_posts_sheet[f'C{row_idx}'] = post['total_engagement']
            top_posts_sheet[f'D{row_idx}'] = post['likes']
            top_posts_sheet[f'E{row_idx}'] = post['comments']
            top_posts_sheet[f'F{row_idx}'] = post['shares']
            top_posts_sheet[f'G{row_idx}'] = post['restacks']
            top_posts_sheet[f'H{row_idx}'] = post['word_count']
            top_posts_sheet[f'I{row_idx}'] = post['reading_time']
            top_posts_sheet[f'J{row_idx}'] = post['pub_date']
        
        # Auto-adjust column widths
        for sheet in [overview_sheet, analytics_sheet, posts_sheet, top_posts_sheet]:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(filename)
        return filename

# =============================================================================
# CONFIGURATION: Change this URL to analyze any Substack publication
# =============================================================================
# Simply replace the URL below with any Substack publication you want to analyze
# Examples:
# - "https://cashandcache.substack.com" (full URL)
# - "https://example.substack.com" (any publication)
# - "example" (just the publication name)
# =================================================================================================

PUBLICATION_URL = "https://cashandcache.com"  # ← CHANGE THIS URL

# =================================================================================================
# Example usage
if __name__ == "__main__":
    # Create collector with the configured URL
    collector = SubstackDataCollector(PUBLICATION_URL)
    
    print(">>> Starting Substack Analytics Dashboard")
    print("=" * 60)
    print(f"Analyzing: {collector.base_url}")
    print(f"Publication: {collector.publication_name}")
    print("=" * 60)
    
    # Run comprehensive analysis - analyze ALL posts
    analysis = collector.analyze_publication()
    
    if "error" in analysis:
        print(f"ERROR: {analysis['error']}")
    else:
        # Display publication overview
        pub = analysis['publication']
        print(f"\n[PUBLICATION OVERVIEW]")
        print(f"Name: {pub['name']}")
        print(f"URL: {pub['url']}")
        print(f"Subscribers: {pub['subscriber_count'] or 'Not publicly available (privacy setting)'}")
        print(f"Last Updated: {pub['last_updated']}")
        
        # Display analytics
        analytics = analysis['analytics']
        print(f"\n[PERFORMANCE METRICS]")
        print(f"Posts Analyzed: {analytics['total_posts_analyzed']}")
        print(f"Average Likes per Post: {analytics['average_likes_per_post']}")
        print(f"Average Comments per Post: {analytics['average_comments_per_post']}")
        print(f"Average Shares per Post: {analytics['average_shares_per_post']}")
        print(f"Average Restacks per Post: {analytics['average_restacks_per_post']}")
        print(f"Average Word Count: {analytics['average_word_count']}")
        print(f"Average Reading Time: {analytics['average_reading_time']} minutes")
        print(f"Publishing Frequency: {analytics['publishing_frequency']} posts/week")
        print(f"Total Engagement: {analytics['total_engagement']}")
        
        # Display top posts
        print(f"\n[TOP PERFORMING POSTS]")
        for i, post in enumerate(analysis['top_posts'][:5], 1):
            title = post['title'].encode('ascii', 'ignore').decode('ascii')
            print(f"{i}. {title}")
            print(f"   Engagement: {post['total_engagement']} (Likes: {post['likes']}, Comments: {post['comments']}, Shares: {post['shares']}, Restacks: {post['restacks']})")
            print(f"   Word Count: {post['word_count']}, Reading Time: {post['reading_time']} min")
            print(f"   Published: {post['pub_date']}")
            print("-" * 50)
        
        print(f"\n[ANALYSIS COMPLETE] Check the detailed data above.")
        
        # Export to Excel
        print(f"\n[EXPORT] Creating Excel file...")
        excel_filename = collector.export_to_excel(analysis)
        print(f"[EXPORT] Excel file created: {excel_filename}")
        print(f"[EXPORT] File saved in current directory: {excel_filename}")
